import asyncio
import logging
import re
import unicodedata
from dataclasses import dataclass, field
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font
from fastapi import HTTPException, status
from sqlalchemy import select, func, or_, and_, cast, update, bindparam
from sqlalchemy.dialects.postgresql import insert as pg_insert, JSONB, array as pg_array
from sqlalchemy.ext.asyncio import AsyncSession
from app.models.product import Product
from app.models.taxonomy import Category, product_categories
from app.models.vehicle import Vehicle, product_vehicles
from app.models.attribute import Attribute, VariantGroup
from app.services.attribute_service import coerce_and_validate_value

logger = logging.getLogger(__name__)

MAX_ROWS_PER_SHEET = 20_000
MAX_FILE_SIZE_BYTES = 4 * 1024 * 1024  # 4 MB - limite crudo antes de siquiera abrir el workbook
MAX_VEHICLE_COMBOS_PER_QUERY = 300  # troceo defensivo, ver _resolve_vehicle_combos

CATEGORIES_SHEET = "Categorias"
VEHICLES_SHEET = "Vehiculos"
ATTRIBUTES_SHEET = "Atributos"
VARIANTS_SHEET = "Variantes"
CATEGORIES_REQUIRED_COLUMNS = ("sku", "categorySlug")
VEHICLES_REQUIRED_COLUMNS = ("sku", "make", "model", "year")  # vehicleType/engine son opcionales
ATTRIBUTES_REQUIRED_COLUMNS = ("sku", "attributeSlug", "value")
VARIANTS_REQUIRED_COLUMNS = ("sku", "variantGroupSlug")


@dataclass
class _RowError:
    row: int
    reason_code: str
    reason: str
    sku: str | None = None


@dataclass
class _SheetOutcome:
    found: bool = False
    processed_rows: int = 0
    assigned_count: int = 0
    errors: list[_RowError] = field(default_factory=list)


@dataclass
class BulkImportOutcome:
    categories: _SheetOutcome
    vehicles: _SheetOutcome
    attributes: _SheetOutcome
    variants: _SheetOutcome


def _load_workbook(file_bytes: bytes):
    """Un archivo invalido/corrupto lanza distintas excepciones segun el caso - todas se
    capturan como un 400 uniforme."""
    try:
        return load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        logger.warning(f"Bulk import: archivo invalido/corrupto ({exc}).")
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Archivo invalido o corrupto - se espera un .xlsx valido.")


def _read_sheet(wb, sheet_name: str, required_columns: tuple[str, ...]) -> tuple[bool, list[tuple[int, dict]]]:
    """Devuelve (existe_la_hoja, filas), fila = (numero_de_fila_excel, dict columna->valor).
    Falta de la hoja no es error (puede traerse solo una de las dos); hoja presente sin
    columnas requeridas si lo es (422, archivo completo). Limite de filas se cuenta
    iterando, no via ws.max_row (impreciso en modo read_only en archivos editados a mano)."""
    if sheet_name not in wb.sheetnames:
        return False, []
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return True, []  # hoja presente pero completamente vacia (ni header)

    header_map = {str(h).strip(): idx for idx, h in enumerate(header) if h is not None}
    missing = [c for c in required_columns if c not in header_map]
    if missing:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"Hoja '{sheet_name}': faltan columnas requeridas: {', '.join(missing)}.",
        )

    rows: list[tuple[int, dict]] = []
    for excel_row_num, raw_row in enumerate(rows_iter, start=2):
        if all(v is None for v in raw_row):
            continue  # fila completamente vacia - se ignora, no cuenta para el limite ni se reporta
        if len(rows) >= MAX_ROWS_PER_SHEET:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Hoja '{sheet_name}': excede el limite de {MAX_ROWS_PER_SHEET} filas de datos.",
            )
        row_dict = {col: raw_row[idx] if idx < len(raw_row) else None for col, idx in header_map.items()}
        rows.append((excel_row_num, row_dict))
    return True, rows


def _parse_workbook(file_bytes: bytes):
    """Carga el .xlsx y lee las 4 hojas de una sola vez - openpyxl (zip+XML) es
    sincrono/bloqueante, asi que import_bulk_assignments corre esto entero en un thread
    (asyncio.to_thread) en vez de bloquear el event loop, mismo criterio que
    google_auth_service.py ya usa para PyJWKClient.fetch_data()."""
    wb = _load_workbook(file_bytes)
    try:
        return (
            _read_sheet(wb, CATEGORIES_SHEET, CATEGORIES_REQUIRED_COLUMNS),
            _read_sheet(wb, VEHICLES_SHEET, VEHICLES_REQUIRED_COLUMNS),
            _read_sheet(wb, ATTRIBUTES_SHEET, ATTRIBUTES_REQUIRED_COLUMNS),
            _read_sheet(wb, VARIANTS_SHEET, VARIANTS_REQUIRED_COLUMNS),
        )
    finally:
        wb.close()


def _clean_str(value) -> str:
    return str(value).strip() if value is not None else ""


def _split_slugs(value) -> list[str]:
    """`categorySlug` acepta uno o varios slugs separados por coma/punto y coma en la misma
    celda; sin delimitador se devuelve como lista de un solo elemento (compatible con el
    formato de una categoria por fila)."""
    raw = _clean_str(value)
    if not raw:
        return []
    return [token.strip() for token in re.split(r"[,;]", raw) if token.strip()]


async def _resolve_products(db: AsyncSession, skus: set[str]) -> dict[str, int]:
    """sku (case-insensitive) -> Product.id, solo productos activos. Resuelve primero por
    Product.sku exacto y luego, para los que quedan sin resolver, hace fallback contra
    additional_skus (JSON, casteado a JSONB inline)."""
    if not skus:
        return {}

    upper_skus = {s.upper() for s in skus}
    result = await db.execute(
        select(Product.id, Product.sku).where(func.upper(Product.sku).in_(upper_skus), Product.is_deleted == False)
    )
    found = {row.sku.upper(): row.id for row in result.all()}

    remaining = upper_skus - found.keys()
    if remaining:
        fallback = await db.execute(
            select(Product.id, Product.additional_skus).where(
                Product.is_deleted == False,
                Product.additional_skus.isnot(None),
                cast(Product.additional_skus, JSONB).has_any(pg_array(sorted(remaining))),
            )
        )
        for row in fallback.all():
            for candidate_sku in (row.additional_skus or []):
                upper_candidate = candidate_sku.upper()
                if upper_candidate in remaining and upper_candidate not in found:
                    found[upper_candidate] = row.id
    return found


async def _resolve_categories(db: AsyncSession, slugs: set[str]) -> dict[str, str]:
    """slug (case-sensitive - ya vienen normalizados en minuscula por el slugify existente) ->
    Category.uuid, un solo IN (...) (tabla chica, decenas de filas)."""
    if not slugs:
        return {}
    result = await db.execute(select(Category.uuid, Category.slug).where(Category.slug.in_(slugs)))
    return {row.slug: row.uuid for row in result.all()}


async def _resolve_attributes(db: AsyncSession, slugs: set[str]) -> dict[str, Attribute]:
    """slug -> Attribute (objeto completo, no solo uuid - se necesita data_type/
    allowed_values para validar cada valor), un solo IN (...) (tabla chica)."""
    if not slugs:
        return {}
    result = await db.execute(select(Attribute).where(Attribute.slug.in_(slugs)))
    return {a.slug: a for a in result.scalars().all()}


async def _resolve_variant_groups(db: AsyncSession, slugs: set[str]) -> dict[str, str]:
    """slug del NOMBRE del grupo, slugificado igual que categorias/atributos, -> VariantGroup.uuid.
    `VariantGroup` no tiene columna `slug` propia (no se expone por URL como categories/
    attributes) - se compara contra una version slugificada de `name` en Python, tabla chica."""
    if not slugs:
        return {}
    result = await db.execute(select(VariantGroup.uuid, VariantGroup.name))
    return {_slugify_ascii(name): uuid for uuid, name in result.all() if _slugify_ascii(name) in slugs}


def _slugify_ascii(name: str) -> str:
    """Mismo slugify que taxonomy_service._slugify/attribute_service._slugify, duplicado
    aqui (sin fallback especifico) solo para comparar `variantGroupSlug` contra `name` -
    VariantGroup no persiste su propio slug."""
    ascii_name = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", "-", ascii_name.lower()).strip("-")


async def _resolve_vehicle_combos(
    db: AsyncSession, combos: set[tuple[str | None, str, str, str | None]]
) -> dict[tuple[str | None, str, str, str | None], list[Vehicle]]:
    """combo = (vehicleType|None, make, model, engine|None), comparado case-insensitive ->
    todos los fitments que coinciden para cualquier anio (el filtro por anio se aplica
    despues, por fila, en _vehicles_matching_year). Una sola consulta (OR de ANDs) para
    toda la hoja, troceada en bloques de MAX_VEHICLE_COMBOS_PER_QUERY solo como
    salvaguarda."""
    if not combos:
        return {}

    combos_list = sorted(combos, key=lambda c: (c[1], c[2], c[0] or "", c[3] or ""))
    all_vehicles: list[Vehicle] = []
    for i in range(0, len(combos_list), MAX_VEHICLE_COMBOS_PER_QUERY):
        chunk = combos_list[i : i + MAX_VEHICLE_COMBOS_PER_QUERY]
        conditions = []
        for vt, make, model, engine in chunk:
            cond = and_(func.upper(Vehicle.make) == make.upper(), func.upper(Vehicle.model) == model.upper())
            if vt:
                cond = and_(cond, Vehicle.vehicle_type == vt)
            if engine:
                cond = and_(cond, func.upper(Vehicle.engine) == engine.upper())
            conditions.append(cond)
        result = await db.execute(select(Vehicle).where(or_(*conditions)))
        all_vehicles.extend(result.scalars().all())

    # Re-agrupa en memoria porque el OR compartido no indica de que combo vino cada fila.
    by_combo: dict[tuple, list[Vehicle]] = {c: [] for c in combos}
    for v in all_vehicles:
        for vt, make, model, engine in combos:
            if (
                v.make.upper() == make.upper()
                and v.model.upper() == model.upper()
                and (not vt or v.vehicle_type == vt)
                and (not engine or (v.engine or "").upper() == engine.upper())
            ):
                by_combo[(vt, make, model, engine)].append(v)
    return by_combo


def _vehicles_matching_year(vehicles: list[Vehicle], year: int) -> list[Vehicle]:
    return [v for v in vehicles if v.year_start <= year and (v.year_end is None or v.year_end >= year)]


async def _bulk_insert_pairs(db: AsyncSession, table, pairs: list[dict], conflict_cols: list[str], returning_col) -> int:
    """ON CONFLICT DO NOTHING hace esto aditivo e idempotente - subir el mismo archivo dos
    veces no duplica ni falla."""
    if not pairs:
        return 0
    stmt = pg_insert(table).values(pairs).on_conflict_do_nothing(index_elements=conflict_cols).returning(returning_col)
    result = await db.execute(stmt)
    return len(result.all())


def _process_categories_rows(
    rows: list[tuple[int, dict]], product_map: dict[str, int], category_map: dict[str, str]
) -> tuple[list[dict], list[_RowError]]:
    """Cada slug de `categorySlug` se resuelve de forma independiente - un slug invalido no
    bloquea los demas de la misma fila."""
    pairs, errors = [], []
    seen = set()
    for row_num, data in rows:
        sku = _clean_str(data.get("sku"))
        slugs = _split_slugs(data.get("categorySlug"))
        if not sku or not slugs:
            errors.append(_RowError(row_num, "MISSING_FIELDS", "Fila incompleta: falta sku o categorySlug.", sku or None))
            continue
        product_id = product_map.get(sku.upper())
        if product_id is None:
            errors.append(_RowError(row_num, "SKU_NOT_FOUND", f"SKU no encontrado: {sku}.", sku))
            continue
        for slug in slugs:
            category_uuid = category_map.get(slug)
            if category_uuid is None:
                errors.append(_RowError(row_num, "CATEGORY_SLUG_NOT_FOUND", f"Categoria con slug no encontrado: {slug}.", sku))
                continue
            key = (category_uuid, product_id)
            if key not in seen:
                seen.add(key)
                pairs.append({"category_uuid": category_uuid, "product_id": product_id})
    return pairs, errors


def _process_vehicles_rows(
    rows: list[tuple[int, dict]],
    product_map: dict[str, int],
    combo_map: dict[tuple, list[Vehicle]],
) -> tuple[list[dict], list[_RowError]]:
    pairs, errors = [], []
    seen = set()
    for row_num, data in rows:
        sku = _clean_str(data.get("sku"))
        make = _clean_str(data.get("make"))
        model = _clean_str(data.get("model"))
        vehicle_type = _clean_str(data.get("vehicleType")) or None
        engine = _clean_str(data.get("engine")) or None
        year_raw = data.get("year")

        if not sku or not make or not model or year_raw is None:
            errors.append(_RowError(row_num, "MISSING_FIELDS", "Fila incompleta: falta sku, make, model o year.", sku or None))
            continue
        try:
            year = int(year_raw)
        except (TypeError, ValueError):
            errors.append(_RowError(row_num, "INVALID_YEAR", f"year invalido (no es un entero): {year_raw!r}.", sku))
            continue

        product_id = product_map.get(sku.upper())
        if product_id is None:
            errors.append(_RowError(row_num, "SKU_NOT_FOUND", f"SKU no encontrado: {sku}.", sku))
            continue

        candidates = combo_map.get((vehicle_type, make, model, engine), [])
        matches = _vehicles_matching_year(candidates, year)
        if not matches:
            suffix = f" motor={engine}" if engine else ""
            errors.append(_RowError(
                row_num, "VEHICLE_NOT_FOUND",
                f"No se encontro ningun vehiculo para {make} {model} {year}{suffix}.", sku,
            ))
            continue

        for v in matches:
            key = (v.uuid, product_id)
            if key not in seen:
                seen.add(key)
                pairs.append({"vehicle_uuid": v.uuid, "product_id": product_id})
    return pairs, errors


def _coerce_cell_value(attribute: Attribute, raw):
    """Celdas de Excel no vienen tipadas como JSON - normaliza el valor crudo al tipo
    Python que `attribute_service.coerce_and_validate_value` espera antes de validarlo.
    Lanza ValueError (capturado por el caller, se reporta como VALUE_TYPE_MISMATCH)."""
    if raw is None:
        return None
    if attribute.data_type == "BOOLEAN":
        if isinstance(raw, bool):
            return raw
        text = _clean_str(raw).lower()
        if text in ("true", "1", "si", "sí", "verdadero", "x"):
            return True
        if text in ("false", "0", "no", "falso", ""):
            return False
        raise ValueError(f"valor booleano no reconocido: {raw!r}")
    if attribute.data_type == "NUMBER":
        if isinstance(raw, bool):
            raise ValueError(f"valor numerico invalido: {raw!r}")
        if isinstance(raw, (int, float)):
            return raw
        try:
            return float(_clean_str(raw))
        except ValueError:
            raise ValueError(f"valor numerico invalido: {raw!r}")
    return _clean_str(raw)  # TEXT / ENUM


def _process_attributes_rows(
    rows: list[tuple[int, dict]], product_map: dict[str, int], attribute_map: dict[str, Attribute]
) -> tuple[dict[int, dict[str, object]], list[_RowError]]:
    """Devuelve product_id -> {slug: value} (varias filas del mismo sku se acumulan en el
    mismo dict; sku+attributeSlug repetido dentro del archivo, la ultima fila gana). No es
    ON CONFLICT DO NOTHING como Categorias/Vehiculos: esto lleva un valor real, un
    reintento con un valor corregido SI debe aplicarse - ver _apply_attribute_updates."""
    updates: dict[int, dict[str, object]] = {}
    errors: list[_RowError] = []
    for row_num, data in rows:
        sku = _clean_str(data.get("sku"))
        attribute_slug = _clean_str(data.get("attributeSlug"))
        if not sku or not attribute_slug:
            errors.append(_RowError(row_num, "MISSING_FIELDS", "Fila incompleta: falta sku o attributeSlug.", sku or None))
            continue
        product_id = product_map.get(sku.upper())
        if product_id is None:
            errors.append(_RowError(row_num, "SKU_NOT_FOUND", f"SKU no encontrado: {sku}.", sku))
            continue
        attribute = attribute_map.get(attribute_slug)
        if attribute is None:
            errors.append(_RowError(row_num, "ATTRIBUTE_SLUG_NOT_FOUND", f"Atributo con slug no encontrado: {attribute_slug}.", sku))
            continue
        try:
            coerced = _coerce_cell_value(attribute, data.get("value"))
            value = coerce_and_validate_value(attribute, coerced)
        except ValueError as exc:
            errors.append(_RowError(row_num, "VALUE_TYPE_MISMATCH", str(exc), sku))
            continue
        updates.setdefault(product_id, {})[attribute.slug] = value
    return updates, errors


def _process_variantes_rows(
    rows: list[tuple[int, dict]], product_map: dict[str, int], variant_group_map: dict[str, str]
) -> tuple[dict[int, str], list[_RowError]]:
    """Devuelve product_id -> VariantGroup.uuid. REEMPLAZO (no aditivo, a diferencia de
    Categorias/Vehiculos): variant_group_uuid es un solo valor por producto, no un tag -
    sku repetido en el archivo, la ultima fila gana."""
    updates: dict[int, str] = {}
    errors: list[_RowError] = []
    for row_num, data in rows:
        sku = _clean_str(data.get("sku"))
        variant_group_slug = _clean_str(data.get("variantGroupSlug"))
        if not sku or not variant_group_slug:
            errors.append(_RowError(row_num, "MISSING_FIELDS", "Fila incompleta: falta sku o variantGroupSlug.", sku or None))
            continue
        product_id = product_map.get(sku.upper())
        if product_id is None:
            errors.append(_RowError(row_num, "SKU_NOT_FOUND", f"SKU no encontrado: {sku}.", sku))
            continue
        variant_group_uuid = variant_group_map.get(variant_group_slug)
        if variant_group_uuid is None:
            errors.append(_RowError(row_num, "VARIANT_GROUP_SLUG_NOT_FOUND", f"Grupo de variantes con slug no encontrado: {variant_group_slug}.", sku))
            continue
        updates[product_id] = variant_group_uuid
    return updates, errors


async def _apply_attribute_updates(db: AsyncSession, updates: dict[int, dict[str, object]]) -> int:
    """Merge (no reemplazo) por producto: lee Product.attributes actual y lo combina con
    las claves nuevas antes de escribir, para que una corrida no borre atributos que otra
    hoja/corrida anterior ya habia guardado. Devuelve pares (producto, atributo)
    realmente escritos (cuenta TODAS las claves aplicadas, no solo las nuevas - un valor
    corregido para una clave existente tambien cuenta)."""
    if not updates:
        return 0
    product_ids = list(updates.keys())
    result = await db.execute(select(Product.id, Product.attributes).where(Product.id.in_(product_ids)))
    current = {row.id: (row.attributes or {}) for row in result.all()}

    rows_to_update = []
    written = 0
    for product_id, new_values in updates.items():
        merged = {**current.get(product_id, {}), **new_values}
        rows_to_update.append({"pid": product_id, "attrs": merged})
        written += len(new_values)

    stmt = update(Product).where(Product.id == bindparam("pid")).values(attributes=bindparam("attrs", type_=JSONB))
    await db.execute(stmt, rows_to_update)
    return written


async def _apply_variant_group_updates(db: AsyncSession, updates: dict[int, str]) -> int:
    if not updates:
        return 0
    stmt = update(Product).where(Product.id == bindparam("pid")).values(variant_group_uuid=bindparam("vgu"))
    await db.execute(stmt, [{"pid": pid, "vgu": vgu} for pid, vgu in updates.items()])
    return len(updates)


async def import_bulk_assignments(db: AsyncSession, file_bytes: bytes) -> BulkImportOutcome:
    """Orquestador: parsea el .xlsx, resuelve productos/categorias/vehiculos/atributos/
    grupos de variantes en un numero constante de consultas, y hace un solo commit final.
    Categorias/Vehiculos son ADITIVOS (ON CONFLICT DO NOTHING); Atributos hace merge
    (no pisa claves de una corrida anterior) y Variantes REEMPLAZA (variant_group_uuid es
    un solo valor por producto, no un tag) - ver los docstrings de cada _process_*_rows.
    Las validaciones por fila nunca tocan la base de datos."""
    if len(file_bytes) > MAX_FILE_SIZE_BYTES:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Archivo demasiado grande (limite {MAX_FILE_SIZE_BYTES // (1024 * 1024)} MB).",
        )

    (cat_found, cat_rows), (veh_found, veh_rows), (attr_found, attr_rows), (var_found, var_rows) = (
        await asyncio.to_thread(_parse_workbook, file_bytes)
    )

    if not cat_found and not veh_found and not attr_found and not var_found:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"El archivo no contiene ninguna hoja '{CATEGORIES_SHEET}', '{VEHICLES_SHEET}', '{ATTRIBUTES_SHEET}' ni '{VARIANTS_SHEET}'.",
        )

    all_skus = (
        {_clean_str(r.get("sku")) for _, r in cat_rows if _clean_str(r.get("sku"))}
        | {_clean_str(r.get("sku")) for _, r in veh_rows if _clean_str(r.get("sku"))}
        | {_clean_str(r.get("sku")) for _, r in attr_rows if _clean_str(r.get("sku"))}
        | {_clean_str(r.get("sku")) for _, r in var_rows if _clean_str(r.get("sku"))}
    )
    product_map = await _resolve_products(db, all_skus)

    category_slugs: set[str] = set()
    for _, r in cat_rows:
        category_slugs.update(_split_slugs(r.get("categorySlug")))
    category_map = await _resolve_categories(db, category_slugs)

    vehicle_combos: set[tuple[str | None, str, str, str | None]] = set()
    for _, r in veh_rows:
        make = _clean_str(r.get("make"))
        model = _clean_str(r.get("model"))
        if make and model:
            vt = _clean_str(r.get("vehicleType")) or None
            eng = _clean_str(r.get("engine")) or None
            vehicle_combos.add((vt, make, model, eng))
    combo_map = await _resolve_vehicle_combos(db, vehicle_combos)

    attribute_slugs = {_clean_str(r.get("attributeSlug")) for _, r in attr_rows if _clean_str(r.get("attributeSlug"))}
    attribute_map = await _resolve_attributes(db, attribute_slugs)

    variant_group_slugs = {_slugify_ascii(_clean_str(r.get("variantGroupSlug"))) for _, r in var_rows if _clean_str(r.get("variantGroupSlug"))}
    variant_group_map = await _resolve_variant_groups(db, variant_group_slugs)

    cat_pairs, cat_errors = _process_categories_rows(cat_rows, product_map, category_map)
    veh_pairs, veh_errors = _process_vehicles_rows(veh_rows, product_map, combo_map)
    attr_updates, attr_errors = _process_attributes_rows(attr_rows, product_map, attribute_map)
    var_updates, var_errors = _process_variantes_rows(var_rows, product_map, variant_group_map)

    cat_assigned = await _bulk_insert_pairs(db, product_categories, cat_pairs, ["category_uuid", "product_id"], product_categories.c.category_uuid)
    veh_assigned = await _bulk_insert_pairs(db, product_vehicles, veh_pairs, ["vehicle_uuid", "product_id"], product_vehicles.c.vehicle_uuid)
    attr_assigned = await _apply_attribute_updates(db, attr_updates)
    var_assigned = await _apply_variant_group_updates(db, var_updates)
    await db.commit()

    logger.info(
        f"Bulk import via /admin: Categorias={len(cat_rows)} filas/{cat_assigned} nuevos/{len(cat_errors)} errores, "
        f"Vehiculos={len(veh_rows)} filas/{veh_assigned} nuevos/{len(veh_errors)} errores, "
        f"Atributos={len(attr_rows)} filas/{attr_assigned} escritos/{len(attr_errors)} errores, "
        f"Variantes={len(var_rows)} filas/{var_assigned} productos/{len(var_errors)} errores."
    )
    return BulkImportOutcome(
        categories=_SheetOutcome(cat_found, len(cat_rows), cat_assigned, cat_errors),
        vehicles=_SheetOutcome(veh_found, len(veh_rows), veh_assigned, veh_errors),
        attributes=_SheetOutcome(attr_found, len(attr_rows), attr_assigned, attr_errors),
        variants=_SheetOutcome(var_found, len(var_rows), var_assigned, var_errors),
    )


def _write_header(ws, columns: tuple[str, ...], comments: dict[str, str]) -> None:
    """Reusa exactamente los nombres de columna que `_read_sheet` espera, para que la
    plantilla nunca se desalinee del lado que la lee."""
    for idx, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx, value=name)
        cell.font = Font(bold=True)
        if name in comments:
            cell.comment = Comment(comments[name], "API SICARX")
        ws.column_dimensions[cell.column_letter].width = max(14, len(name) + 4)


def build_template_workbook() -> bytes:
    """Mismas hojas/columnas que `_read_sheet` exige. Filas de ejemplo con datos ficticios -
    si un admin la sube sin editarla, cada fila sale como error por fila, no fatal (el
    diseno de exito parcial de import_bulk_assignments ya cubre ese caso)."""
    wb = Workbook()

    ws_cat = wb.active
    ws_cat.title = CATEGORIES_SHEET
    _write_header(ws_cat, CATEGORIES_REQUIRED_COLUMNS, {
        "categorySlug": "Slug (no nombre ni uuid) de una categoria ya existente - ver GET /taxonomy o GET /admin/categories. Para asignar varias categorias al mismo producto en una sola fila, separalas por coma o punto y coma (p. ej. \"categoria-ejemplo-1, categoria-ejemplo-2\").",
    })
    ws_cat.append(["SKU-EJEMPLO-1", "categoria-ejemplo-1"])
    ws_cat.append(["SKU-EJEMPLO-1", "categoria-ejemplo-2, categoria-ejemplo-3"])

    ws_veh = wb.create_sheet(VEHICLES_SHEET)
    veh_columns = VEHICLES_REQUIRED_COLUMNS + ("vehicleType", "engine")
    _write_header(ws_veh, veh_columns, {
        "make": "No distingue mayusculas/minusculas (igual que model, engine y sku).",
        "model": "No distingue mayusculas/minusculas (igual que make, engine y sku).",
        "year": "Un solo anio (no un rango yearStart/yearEnd) - se compara contra el rango de los fitments ya existentes en 'vehicles'.",
        "vehicleType": "Opcional: AUTOMOTIVE o MOTORCYCLE. Dejar vacio si no hace falta desambiguar.",
        "engine": "Opcional - dejar vacio para aplicar a TODAS las variantes de motor de esa marca/modelo/anio.",
    })
    # Orden de columnas: sku, make, model, year, vehicleType, engine (ver veh_columns arriba)
    ws_veh.append(["SKU-EJEMPLO-2", "Marca-Ejemplo", "Modelo-Ejemplo", 2020, None, None])
    ws_veh.append(["SKU-EJEMPLO-2", "Marca-Ejemplo", "Modelo-Ejemplo", 2020, "AUTOMOTIVE", "L4 1.6L"])

    ws_attr = wb.create_sheet(ATTRIBUTES_SHEET)
    _write_header(ws_attr, ATTRIBUTES_REQUIRED_COLUMNS, {
        "attributeSlug": "Slug (no nombre ni uuid) de un atributo ya existente - ver GET /admin/attributes. Una fila por (sku, atributo); para varios atributos del mismo producto, repite el sku en varias filas.",
        "value": "Formato segun el dataType del atributo: TEXT/ENUM = texto tal cual (ENUM debe ser uno de sus allowedValues); NUMBER = numero; BOOLEAN = TRUE/FALSE (tambien acepta si/no, 1/0). A diferencia de Categorias/Vehiculos, MERGE - un valor corregido en una corrida posterior SI se aplica, no se ignora.",
    })
    ws_attr.append(["SKU-EJEMPLO-1", "atributo-ejemplo-color", "Rojo"])
    ws_attr.append(["SKU-EJEMPLO-1", "atributo-ejemplo-voltaje", 12])

    ws_var = wb.create_sheet(VARIANTS_SHEET)
    _write_header(ws_var, VARIANTS_REQUIRED_COLUMNS, {
        "variantGroupSlug": "Slug derivado del NOMBRE de un grupo de variantes ya existente - ver GET /admin/variant-groups (VariantGroup no tiene slug propio, se deriva de su name igual que categorias/atributos). A diferencia de Categorias/Vehiculos, REEMPLAZA: variantGroupUuid es un solo valor por producto, no un tag.",
    })
    ws_var.append(["SKU-EJEMPLO-1", "grupo-de-variantes-ejemplo"])
    ws_var.append(["SKU-EJEMPLO-3", "grupo-de-variantes-ejemplo"])

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
